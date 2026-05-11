import openpyxl
import pandas as pd

from utilities.read_configs_properties import ReadConfig


# from openpyxl.styles import PatternFill

def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_row)

def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def read_data(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.cell(row_num, column_num).value)

def write_data(file, sheetname, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    # Assign data to the specific cell
    sheet.cell(row=row_num, column=column_num).value = data
    # Save the workbook to apply changes
    workbook.save(file)

def read_data_pandas():
    admin_login_file = ReadConfig.get_admin_login_file()
    data = pd.read_excel(admin_login_file, sheet_name="Sheet1", engine="openpyxl")
    return data

def write_data_pandas(df):
    admin_login_file = ReadConfig.get_admin_login_file()
    df.to_excel(admin_login_file, index=False)


def get_customers_data():
    df = pd.read_excel(
        ReadConfig.get_customers_data_file(),
        sheet_name="Sheet1",
        engine="openpyxl"
    )
    return df.to_dict(orient="records") # converted to dictionary via to_dict. each record is a dict. For ex: data = {"email": "test_admin@yourstore.com","password": "123",}